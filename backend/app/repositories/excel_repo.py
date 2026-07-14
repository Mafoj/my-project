"""
Phase 1 repository: reads the Excel export that the PMO maintains by hand.

This replaces the browser-side XLSX.js parsing. Same column maps, same coercion
rules -- just executed once on the server instead of once per user per page load.

Caching: the parsed payload is held in memory and invalidated on file mtime
change. A PMO user dropping a new export into data/ is picked up on the next
request; no restart needed.
"""
from __future__ import annotations

import logging
import threading
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.coerce import coerce_date, nk, norm_probability, norm_status, to_num
from app.models import MONTH_KEYS, Allocation, DatasetMeta, PipelinePayload, Project
from app.repositories.base import PipelineRepository

log = logging.getLogger(__name__)

# Ported verbatim from MAIN_MAP in the original app. Keys are nk()-normalised
# spreadsheet headers; values are our field names. Multiple spellings map to the
# same field on purpose -- the source spreadsheet is inconsistent.
MAIN_MAP: dict[str, str] = {
    "project name": "project_name",
    "project name (as in its1)": "project_name_its",
    "tower": "tower",
    "pc ownership": "pc_ownership",
    "pm name": "pm_name",
    "internal/ external": "int_ext",
    "internal/external": "int_ext",
    "salesforce proposal # / bca #": "sales_force",
    "#bso/io": "bso_io",
    "project status": "project_status",
    "start date": "start_date",
    "end date": "end_date",
    "bu": "bu",
    "value 2026": "value_2026",
    "probability [%]": "probability",
    "value weighted 2026": "value_weighted_2026",
    "comments": "comments",
}

# Ported verbatim from ALLOC_MAP.
ALLOC_MAP: dict[str, str] = {
    "initial ip tower": "tower",
    "profit ownership": "profit_ownership",
    "pm depart": "pm_depart",
    "pm name": "pm_name",
    "pm name secondary": "pm_name_secondary",
    "project name": "project_name",
    "externality": "externality",
    "order n.": "order_n",
    "notes": "notes",
    "value": "value",
    "project status": "project_status",
    "probability": "probability",
    "is it in project list?": "in_project_list",
    **{m: f"month_{m}" for m in MONTH_KEYS},
}


def _parse_sheet(ws: Any, col_map: dict[str, str]) -> list[dict[str, Any]]:
    """Python equivalent of parseSheet(): header row 0, map columns, drop blanks."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    idx: dict[str, int] = {}
    for i, header in enumerate(rows[0]):
        key = nk(header)
        field = col_map.get(key) or col_map.get(key.rstrip())
        if field and field not in idx:
            idx[field] = i

    missing = set(col_map.values()) - set(idx)
    if missing:
        log.warning("Sheet %r: unmapped columns (will default): %s", ws.title, sorted(missing))

    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(c not in (None, "") for c in row):
            continue
        rec: dict[str, Any] = {}
        for field, i in idx.items():
            v = row[i] if i < len(row) else ""
            rec[field] = v.strip() if isinstance(v, str) else v
        out.append(rec)
    return out


class ExcelPipelineRepository(PipelineRepository):
    def __init__(self, path: Path) -> None:
        self._path = path
        self._lock = threading.Lock()
        self._cache: PipelinePayload | None = None
        self._cached_mtime: float | None = None

    def health(self) -> tuple[bool, str]:
        if not self._path.exists():
            return False, f"Excel source not found: {self._path}"
        return True, f"Excel source OK: {self._path.name}"

    def load(self) -> PipelinePayload:
        if not self._path.exists():
            raise FileNotFoundError(f"Pipeline export not found at {self._path}")

        mtime = self._path.stat().st_mtime
        with self._lock:
            if self._cache is not None and self._cached_mtime == mtime:
                return self._cache
            log.info("Parsing pipeline export %s (mtime=%s)", self._path.name, mtime)
            payload = self._parse(mtime)
            self._cache, self._cached_mtime = payload, mtime
            return payload

    def _parse(self, mtime: float) -> PipelinePayload:
        # read_only + data_only: we want computed values, not formulas, and we
        # never write back. This is what makes a large export cheap to parse.
        wb = load_workbook(self._path, read_only=True, data_only=True)
        try:
            main_ws = wb[wb.sheetnames[0]]
            alloc_name = next(
                (n for n in wb.sheetnames if "alloc" in n.lower()), wb.sheetnames[0]
            )
            alloc_ws = wb[alloc_name]

            projects = [self._to_project(r) for r in _parse_sheet(main_ws, MAIN_MAP)]
            allocations = [self._to_alloc(r) for r in _parse_sheet(alloc_ws, ALLOC_MAP)]
        finally:
            wb.close()

        meta = DatasetMeta(
            source="excel",
            source_ref=self._path.name,
            generated_at=datetime.now(UTC).isoformat(),
            source_modified_at=datetime.fromtimestamp(mtime, UTC).isoformat(),
            project_count=len(projects),
            allocation_count=len(allocations),
        )
        log.info("Parsed %d projects, %d allocations", len(projects), len(allocations))
        return PipelinePayload(meta=meta, projects=projects, allocations=allocations)

    @staticmethod
    def _to_project(r: dict[str, Any]) -> Project:
        return Project(
            project_name=str(r.get("project_name") or ""),
            project_name_its=str(r.get("project_name_its") or ""),
            tower=str(r.get("tower") or ""),
            pc_ownership=str(r.get("pc_ownership") or ""),
            pm_name=str(r.get("pm_name") or ""),
            int_ext=str(r.get("int_ext") or ""),
            sales_force=str(r.get("sales_force") or ""),
            bso_io=str(r.get("bso_io") or ""),
            project_status=norm_status(r.get("project_status")),
            start_date=coerce_date(r.get("start_date")),
            end_date=coerce_date(r.get("end_date")),
            bu=str(r.get("bu") or ""),
            comments=str(r.get("comments") or ""),
            value_2026=to_num(r.get("value_2026")),
            value_weighted_2026=to_num(r.get("value_weighted_2026")),
            probability=norm_probability(r.get("probability")),
        )

    @staticmethod
    def _to_alloc(r: dict[str, Any]) -> Allocation:
        return Allocation(
            tower=str(r.get("tower") or ""),
            profit_ownership=str(r.get("profit_ownership") or ""),
            pm_depart=str(r.get("pm_depart") or ""),
            pm_name=str(r.get("pm_name") or ""),
            pm_name_secondary=str(r.get("pm_name_secondary") or ""),
            project_name=str(r.get("project_name") or ""),
            externality=str(r.get("externality") or ""),
            order_n=str(r.get("order_n") or ""),
            notes=str(r.get("notes") or ""),
            project_status=norm_status(r.get("project_status")),
            in_project_list=str(r.get("in_project_list") or ""),
            value=to_num(r.get("value")),
            probability=norm_probability(r.get("probability")),
            months={m: to_num(r.get(f"month_{m}")) for m in MONTH_KEYS},
        )
